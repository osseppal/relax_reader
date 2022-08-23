# -*- coding: utf-8 -*-

'''Gleeble relaxation data combiner.
Reads excel files containing gleeble data and combines them into excel files with constant strain, 
strain rate and temperature.

14.09.2021: v1.0 -> basic functions, two parameter exceling
14.09.2021: v2.0 -> three parameter exceling
20.09.2021: v3.0 -> GUI added, with simple do it all- button
21.09.2021: v4.0 -> manual compress|relax start point modification added, combining process split into different buttons
04.10.2021: v5.0 -> strain rate save mode added, stress-strain curve combination added to separate wb
12.10.2021: v6.0 -> data thinning controls added, fixed bug with <adjust c|r>
27.10.2021: v7.0 -> Perttula-type semi-automatic SRX calculation with GUI added, gleeble_data_handler-module added to separate modules with different tasks
23.11.2021: v8.0 -> Zurob-type SRX and recovery fitting added
28.12.2021: v9.0 -> x-y limits added
'''

software_version = '9-0'

from collections import OrderedDict as orddict
import numpy as np
import math
import os
import sys
import openpyxl as opyx

import matplotlib
matplotlib.use("Agg") #this prevents plt.close(fig) from closing GUI
import matplotlib.pyplot as plt
import matplotlib.figure as mpl_fig
import matplotlib.backends.backend_tkagg as tkagg


import tkinter as tk
import tkinter.ttk as ttk
import tkinter.filedialog as filedialog


import common_exceling_relax as cex
import gleeble_data_handler_relax as gdh

    
def constant_dict_filler(d1, tsd, data_str, vars1, vars2, vars3, vars4, var1_name, var2_name, var3_name, var4_name):
    '''Fills a dict with some constants with inner dicts.
        Returns nothing.'''
    
    for var1 in vars1:
        for var2 in vars2:
            d1_1 = d1['%s_%s' %(var1, var2)] = orddict()
            for var3 in vars3:
                for var4 in vars4:
                    for tsd_key, ts_obj in tsd.items():
                        tsd_var1, tsd_var2, tsd_var3, tsd_var4 = ts_obj.var_d[var1_name], ts_obj.var_d[var2_name], ts_obj.var_d[var3_name], ts_obj.var_d[var4_name]
                        key1_check = np.allclose(tsd_var1, var1)
                        key2_check = np.allclose(tsd_var2, var2)
                        key3_check = np.allclose(tsd_var3, var3)
                        key4_check = np.allclose(tsd_var4, var4)
                        # print(var1_name, var2_name, var3_name)
                        # print(tsd_var1, var1)
                        # print(tsd_var2, var2)
                        # print(tsd_var3, var3)
                        if key1_check and key2_check and key3_check and key4_check:
                            try:
                                time_list = ts_obj.data_d[data_str]['Time(sec)']
                                stress_list = ts_obj.data_d[data_str]['Stress(MPa)']
                            except KeyError:
                                break
                            header = '%s_%s_%s_%s' %(var3_name, var3, var4_name, var4)
                            d1_1_dh = d1_1[header] = orddict()
                            d1_1_dh['Time'] = time_list
                            d1_1_dh['Stress'] = stress_list
                            
                            break
    
    return

def find_ts_sub_d(ts_d, T, strain, SR, t_hold):
    ''' '''
    
    for tsd_key, sub_d in ts_d.items():
        my_T, my_strain, my_SR, my_t_hold = sub_d['nom_T'], sub_d['nom_strain'], sub_d['nom_SR'], sub_d['nom_t_hold']
        T_chk = np.allclose(my_T, T)
        strain_chk = my_strain == strain
        SR_chk = np.allclose(my_SR, SR)
        t_hold_chk = np.allclose(my_t_hold, t_hold)
        if T_chk and strain_chk and SR_chk and t_hold_chk:
            ts_sub_d = sub_d
            break
    
    return tsd_key, ts_sub_d


def const_press_d_filler(d1, tsd, data_str, vars1, vars2, vars3, vars4, var1_name, var2_name, var3_name, var4_name, x_var):
    ''' '''
    
    for var1 in vars1:
        for var2 in vars2:
            d1_1 = d1['%s_%s' %(var1, var2)] = orddict()
            for var3 in vars3:
                for var4 in vars4:
                    for tsd_key, ts_obj in tsd.items():
                        tsd_var1, tsd_var2, tsd_var3, tsd_var4 = ts_obj.var_d[var1_name], ts_obj.var_d[var2_name], ts_obj.var_d[var3_name], ts_obj.var_d[var4_name]
                        key1_check = np.allclose(tsd_var1, var1)
                        key2_check = np.allclose(tsd_var2, var2)
                        key3_check = np.allclose(tsd_var3, var3)
                        key4_check = np.allclose(tsd_var4, var4)
                        if key1_check and key2_check and key3_check and key4_check:
                            try:
                                strains = ts_obj.data_d[data_str]['Strain']
                                stresses = ts_obj.data_d[data_str]['Stress(MPa)']
                                r_start = ts_obj.var_d['r_start']
                                cut_loc = r_start[0]+500
                                strains = strains[:cut_loc]
                                stresses = stresses[:cut_loc]
                            except KeyError:
                                break
                            header = '%s%s_%s%s' %(var3_name, var3, var4_name, var4)
                            d1_1_dh = d1_1[header] = orddict()
                            d1_1_dh['Strain'] = strains
                            d1_1_dh['Stress'] = stresses                            
                            break
    
    
    return

def wrap_constant_dict_to_wb(wb_filepath, tsd, data_str, vars1, vars2, vars3, vars4, var1_name, var2_name, var3_name, var4_name, x_var='Time'):
    '''Creates a dict with certain constant values, saves said dict into an xlsx file and creates a chart for each worksheet.'''
    
    print('Creating %s workbook for constant %s with %s...' %(x_var,var1_name, data_str))
    d1 = orddict()
    if x_var == 'Time':
        constant_dict_filler(d1, tsd, data_str, vars1, vars2, vars3, vars4, var1_name, var2_name, var3_name, var4_name)
        # constant_dict_filler(d1, tsd, data_str, vars1, subsub_vars, sub_vars, var_name, subsub_var_name, sub_var_name)
    elif x_var == 'Strain':
        const_press_d_filler(d1, tsd, data_str, vars1, vars2, vars3, vars4, var1_name, var2_name, var3_name, var4_name, x_var)
    wb = cex.opyx_dh_dict_to_wb(wb_filepath, d1)
    print('Hep! creating charts...')
    cex.charts_to_const_wb(wb, x_var)
    # common_column_size(wb, xlsx_filepath, 16.5)
    wb.save(wb_filepath)
    print('Hep!')
    
    return

def common_column_size(wb, xlsx_filepath, col_width):
    '''Automatically sets the column width to best fit the length of labels.'''
    
    ws_names = wb.sheetnames
        
    for this_ws in wb:
        this_max_col = this_ws.max_column
        label_range_gen = this_ws.get_squared_range(1, 1, this_max_col, 1)
        for this_row in label_range_gen:
            for this_label in this_row:
                lab_val = this_label.value
                this_col = this_label.column
                this_ws.column_dimensions[this_col].width = col_width
    
    return

def wrap_load_constant_wbs():
    ''' '''
    
    global tsd
    
    #variables
    temps = list(map(int, dname_d['temps'].get()[1:-1].split(',')))
    SRs = list(map(int, dname_d['SRs'].get()[1:-1].split(',')))
    hold_times = list(map(int, dname_d['hold_times'].get()[1:-1].split(',')))
    strains = list(map(float, dname_d['strains'].get()[1:-1].split(',')))
    #dir paths
    load_dir_pth = dname_d['load_dir_pth'].get()
    save_dir_pth = dname_d['save_dir_pth'].get()
    #cr variables
    c_lim_val = dname_d['c_lim'].get()
    c_forced_time_val = dname_d['c_forced_time'].get()
    r_lim_val = dname_d['r_lim'].get()
    #data handling
    thin_skip = dname_d['thin_skip'].get()
    pavg_range = dname_d['pavg_range'].get()

    tsd = test_series_dict = gdh.wrap_relax_test_series_dicter(load_dir_pth, temps, strains, hold_times, SRs, thin_skip, pavg_range, c_lim_val, c_forced_time_val, r_lim_val, temp_str='', printing=True)
    
    
    return

def save_all_SRX_wb(wb_pth, temps, SRs, strains, hold_times):
    ''' '''
    
    print('Saving all_SRX_wb...')
    wb_d = orddict()
    vars_d = wb_d['variables'] = orddict()
    lbls1 = ['num', 'nom_T', 'nom_strain', 'nom_SR', 'nom_t_hold']
    # perttula_lbls = ['a1', 'b1', 'a2', 'b2', 'k_JMAK', 'n']
    lbls2 = ['a1', 'b1', 'a2', 'b2', 'k_JMAK', 'n', 'U_a', 'V_a', 'Q_d', 'k_Z']
    lbls3 = ['t50_pert', 't50_JMAK', 't50_Zurob', 't50_Zpert']
    lbls = lbls1 + lbls2 + lbls3
    #fill variable dict
    for num, ts_obj in tsd.items():
        for lbl in lbls:
            if lbl == 'num':
                val = num
            else:
                try:
                    val = ts_obj.var_d[lbl]
                except KeyError:
                    try:
                        val = ts_obj.SRX_d[lbl]
                    except KeyError:
                        val = ''
            try:
                vars_d[lbl].append(val)
            except KeyError:
                vars_d[lbl] = [val]
    #add SRX data to each case
    for num, ts_obj in tsd.items():
        try:
            ts_obj.SRX_d['a1']
        except KeyError:
            print('skipping %d' %num)
            continue
        ws_d = wb_d[num] = ts_obj.SRX_data_d

    
    #create workbook
    wb = cex.opyx_dict_to_wb(wb_pth, wb_d)
    
    #save charts
    for sheetname in wb.sheetnames:
        if sheetname == 'variables':
            continue
        ws = wb[sheetname]
        cex.SRX_data_to_little_charts(ws)
    
    wb.save(wb_pth)
    print('%s Finished!' %wb_pth)
    
    return

def wrap_save_constant_wbs():
    ''' '''
    
    #variables
    temps = list(map(int, dname_d['temps'].get()[1:-1].split(',')))
    SRs = list(map(int, dname_d['SRs'].get()[1:-1].split(',')))
    hold_times = list(map(int, dname_d['hold_times'].get()[1:-1].split(',')))
    strains = list(map(float, dname_d['strains'].get()[1:-1].split(',')))
    #dir paths
    # load_dir_pth = dname_d['load_dir_pth'].get()
    save_dir_pth = dname_d['save_dir_pth'].get()
    #radiobuttons
    # test_mode = dname_d['test_mode'].get()
    #checkbuttons
    all_data_pavg_chk = dname_d['all_data_pavg_chk'].get()
    all_data_orig_chk = dname_d['all_data_orig_chk'].get()
    const_temp_chk = dname_d['const_temp_chk'].get()
    const_strain_chk = dname_d['const_strain_chk'].get()
    const_t_hold_chk = dname_d['const_t_hold_chk'].get()
    const_SR_chk = dname_d['const_SR_chk'].get()
    # compr_chk = dname_d['compr_chk'].get()
    # relax_chk = dname_d['relax_chk'].get()
    str_str_chk = dname_d['str-str_chk'].get()
    all_SRX_chk = dname_d['all_SRX_chk'].get()
    ### SRX workbooks
    if all_SRX_chk == 1:
        wb_pth = save_dir_pth + 'all_SRX.xlsx'
        save_all_SRX_wb(wb_pth, temps, SRs, strains, hold_times)
    # return
    ### stress-strain workbooks
    if str_str_chk == 1:
        # wrap_stressstrain_wbs(save_dir_pth, temps, SRs, hold_times, strains)
        wb_pth = save_dir_pth + 'press_const_temp_pavg.xlsx'
        var1_name, var2_name, var3_name, var4_name = 'nom_T', 'nom_SR', 'nom_strain', 'nom_t_hold'
        vars1, vars2, vars3, vars4 = temps, SRs, strains, hold_times
        data_str = 'datas_pavg'
        wrap_constant_dict_to_wb(wb_pth, tsd, data_str, vars1, vars2, vars3, vars4, var1_name, var2_name, var3_name, var4_name, x_var='Strain')
    ### all data workbooks
    log10_chk = 'yes'
    cr_starts = 'yes'
    var_lbls = ['num', 'nom_T', 'nom_SR', 'nom_strain', 'nom_t_hold', 'c_start', 'r_start', 'c_stress', 'r_stress', 'r_strain', 'strain_diff', 'T_min', 'T_max', 'T_diff']
    if all_data_pavg_chk == 1:
        wb_pth = save_dir_pth + 'all_datas_pavg.xlsx'
        cex.wrap_all_data_to_wb(wb_pth, tsd, 'datas_pavg', var_lbls, log10_chk, cr_starts=cr_starts)
    if all_data_orig_chk == 1:
        wb_pth = save_dir_pth + 'all_datas_orig.xlsx'
        cex.wrap_all_data_to_wb(wb_pth, tsd, 'datas', var_lbls, log10_chk, cr_starts=cr_starts)
    ### dividing data into constant workbooks
    # nom_T|nom_SR|nom_strain|nom_t_hold
    ##temperature
    if const_temp_chk == 1:
        var1_name, var2_name, var3_name, var4_name = 'nom_T', 'nom_strain', 'nom_t_hold', 'nom_SR'
        # temp_wb_filepath = save_dir_pth + 
        vars1, vars2, vars3, vars4 = temps, strains, hold_times, SRs
        data_strs = ['datas_pavg', 'datas_compress', 'datas_relax']
        temp_filenames = ['constant_temp_pavg.xlsx', 'constant_temp_C.xlsx', 'constant_temp_R.xlsx']
        for i, data_str in enumerate(data_strs):
            temp_wb_pth = save_dir_pth + temp_filenames[i]
            wrap_constant_dict_to_wb(temp_wb_pth, tsd, data_str, vars1, vars2, vars3, vars4, var1_name, var2_name, var3_name, var4_name)
    ##strain
    if const_strain_chk == 1:
        var1_name, var2_name, var3_name, var4_name = 'nom_strain', 'nom_T', 'nom_t_hold', 'nom_SR'
        # strain_wb_pth = save_dir_pth + 'constant_strain_pavg.xlsx'
        vars1, vars2, vars3, vars4 = strains, temps, hold_times, SRs
        data_strs = ['datas_pavg', 'datas_compress', 'datas_relax']
        strain_filenames = ['constant_strain_pavg.xlsx', 'constant_strain_C.xlsx', 'constant_strain_R.xlsx']
        for i, data_str in enumerate(data_strs):
            strain_wb_pth = save_dir_pth + strain_filenames[i]
            wrap_constant_dict_to_wb(strain_wb_pth, tsd, data_str, vars1, vars2, vars3, vars4, var1_name, var2_name, var3_name, var4_name)
    ##strain rate
    if const_SR_chk == 1:
        var1_name, var2_name, var3_name, var4_name = 'nom_SR', 'nom_T', 'nom_strain', 'nom_t_hold'
        vars1, vars2, vars3, vars4 = SRs, temps, strains, hold_times
        # SR_wb_pth = save_dir_pth + 'constant_SR_pavg.xlsx'
        data_strs = ['datas_pavg', 'datas_compress', 'datas_relax']
        SR_filenames = ['constant_SR_pavg.xlsx', 'constant_SR_C.xlsx', 'constant_SR_R.xlsx']
        for i, data_str in enumerate(data_strs):
            SR_wb_pth = save_dir_pth + SR_filenames[i]        
            wrap_constant_dict_to_wb(SR_wb_pth, tsd, data_str, vars1, vars2, vars3, vars4, var1_name, var2_name, var3_name, var4_name)
    ##t_hold
    if const_t_hold_chk == 1:
        var1_name, var2_name, var3_name, var4_name = 'nom_t_hold', 'nom_T', 'nom_strain', 'nom_SR'
        vars1, vars2, vars3, vars4 = hold_times, temps, strains, SRs
        # SR_wb_pth = save_dir_pth + 'constant_t_hold_pavg.xlsx'
        data_strs = ['datas_pavg', 'datas_compress', 'datas_relax']
        t_hold_filenames = ['constant_t_hold_pavg.xlsx', 'constant_t_hold_C.xlsx', 'constant_t_hold_R.xlsx']
        for i, data_str in enumerate(data_strs):
            t_hold_wb_pth = save_dir_pth + t_hold_filenames[i]        
            wrap_constant_dict_to_wb(t_hold_wb_pth, tsd, data_str, vars1, vars2, vars3, vars4, var1_name, var2_name, var3_name, var4_name)
    print('### Saving sequence complete! ###')
    
    
    return

def create_textboxes(fr):
    ''' '''
    
    ##textboxes
    #dir path
    tbox1_fr = ttk.Labelframe(fr, text='dir paths:')
    tbox1_fr.grid(column=0, row=0, sticky=tk.W, columnspan=30)
    dname_d1 = orddict()
    dname_d1['load_dir_pth'] = tk.StringVar(value=load_dir_pth_init)
    dname_d1['save_dir_pth'] = tk.StringVar(value=save_dir_pth_init)
    wi = 150
    tbox_widgets(tbox1_fr, dname_d1, wi)
    #variable
    tbox2_fr = ttk.Labelframe(fr, text='variables')
    tbox2_fr.grid(column=0, row=1, sticky=tk.W, rowspan=2)
    dname_d2 = orddict()
    dname_d2['temps'] = tk.StringVar(value=temps)
    dname_d2['strains'] = tk.StringVar(value=strains)
    dname_d2['hold_times'] = tk.StringVar(value=hold_times)
    dname_d2['SRs'] = tk.StringVar(value=SRs)
    wi = 50
    tbox_widgets(tbox2_fr, dname_d2, wi)
    #cr variable
    tbox3_fr = ttk.Labelframe(fr, text='c|r variables')
    tbox3_fr.grid(column=2, row=3, sticky=tk.W)
    dname_d3 = orddict()
    dname_d3['c_lim'] = tk.DoubleVar(value=0.05)
    dname_d3['c_forced_time'] = tk.DoubleVar(value=0.061)
    dname_d3['r_lim'] = tk.DoubleVar(value=0.99)
    wi = 10
    tbox_widgets(tbox3_fr, dname_d3, wi)
    #data thinning 
    tbox4_fr = ttk.Labelframe(fr, text='data handling')
    tbox4_fr.grid(column=4, row=3, sticky=tk.W)
    dname_d4 = orddict()
    dname_d4['thin_skip'] = tk.IntVar(value=10)
    dname_d4['pavg_range'] = tk.IntVar(value=15)
    wi = 5
    i = tbox_widgets(tbox4_fr, dname_d4, wi)
    
    ##checkbuttons
    #relax
    cbox1_fr = ttk.Labelframe(fr, text='checks_relax')
    cbox1_fr.grid(column=2, row=1, sticky=tk.W, rowspan=2, columnspan=2)
    dname_d5 = orddict()
    dname_d5['all_data_pavg_chk'] = tk.IntVar(value=1)
    dname_d5['all_data_orig_chk'] = tk.IntVar(value=1)
    dname_d5['const_temp_chk'] = tk.IntVar(value=1)
    dname_d5['const_strain_chk'] = tk.IntVar(value=1)
    dname_d5['const_t_hold_chk'] = tk.IntVar(value=1)
    dname_d5['const_SR_chk'] = tk.IntVar(value=0)
    # dname_d5['compr_chk'] = tk.IntVar(value=1)
    # dname_d5['relax_chk'] = tk.IntVar(value=1)
    cb_widgets(cbox1_fr, dname_d5)
    #press
    cbox2_fr = ttk.Labelframe(fr, text='checks_press')
    cbox2_fr.grid(column=4, row=1, sticky=tk.W)
    dname_d6 = orddict()
    dname_d6['str-str_chk'] = tk.IntVar(value=1)
    cb_widgets(cbox2_fr, dname_d6)
    #SRX
    cbox3_fr = ttk.Labelframe(fr, text='checks_SRX')
    cbox3_fr.grid(column=4, row=2, sticky=tk.W)
    dname_d7 = orddict()
    dname_d7['all_SRX_chk'] = tk.IntVar(value=1)
    cb_widgets(cbox3_fr, dname_d7)
    #combobox
    mat_types = ['1523', '1524', '1525', 'Raex400-old']
    cbox_var = tk.StringVar(value=mat_types[0])
    cbox = ttk.Combobox(fr, textvariable=cbox_var, state='readonly', values=mat_types, width=12)
    cbox.grid(column=3, row=3)
    # cbox.bind('<<ComboboxSelected>>', num_selecter)
    #combine all into master dict
    dname_d = orddict(**dname_d1, **dname_d2, **dname_d3, **dname_d4, **dname_d5, **dname_d6, **dname_d7)
    dname_d['mat_type'] = cbox_var
    
    return dname_d

def rb_widgets(fr, name, var, var_names):
    ''' '''
    
    lf = ttk.Labelframe(fr, text=name)
    lf.grid(column=0, row=0)
    for i, var_name in enumerate(var_names):
        rb = ttk.Radiobutton(lf, text=var_name, variable=var, value=var_name)
        rb.grid(column=0, row=i, sticky=tk.W)
    
    return

def cb_widgets(fr, d1):
    ''' '''
    
    i = 0
    for name, var in d1.items():
        cbtn = ttk.Checkbutton(fr, variable=var, text=name)
        cbtn.grid(column=i//4, row=i%4, sticky=tk.W)
        i += 1
    
    return

def tbox_widgets(fr, d1, wi, col_len=4):
    ''' '''
    
    i = 0
    for name, var in d1.items():
        lbl = tk.Label(fr, text=name)
        tbox = tk.Entry(fr, textvariable=var, width=wi)
        lbl.grid(column=(i//col_len)*2, row=i%col_len, sticky=tk.E)
        tbox.grid(column=(i//col_len)*2+1, row=i%col_len, sticky=tk.W)
        i += 1
    
    return i

def create_buttons(fr):
    ''' '''
    
    bttn_mainfr = tk.Frame(fr)
    bttn_mainfr.grid(column=0, row=3, sticky=tk.W)
    #create buttons
    #-button text variables
    btn_txt_d = orddict()
    btxt_load_dir = btn_txt_d['load_dir'] = tk.StringVar(value='Choose\nload dir')
    btxt_save_dir = btn_txt_d['save_dir'] = tk.StringVar(value='Choose\nsave dir')
    btxt_load = btn_txt_d['load'] = tk.StringVar(value='load\nworkbooks')
    btxt_adjust = btn_txt_d['adjust'] = tk.StringVar(value='adjust\nc|r start')
    btxt_draw = btn_txt_d['draw'] = tk.StringVar(value='Draw\nlines')
    btxt_save = btn_txt_d['save'] = tk.StringVar(value='save\nworkbooks')
    #-buttons
    # bttn_fr = tk.Frame(bttn_mainfr)
    bttn_d = orddict()
    load_dir_bttn = bttn_d['load_dir'] = tk.Button(bttn_mainfr, textvariable=btxt_load_dir, command=load_dir_button)
    save_dir_bttn = bttn_d['save_dir'] = tk.Button(bttn_mainfr, textvariable=btxt_save_dir, command=save_dir_button)
    load_bttn = bttn_d['load'] = tk.Button(bttn_mainfr, textvariable=btxt_load, command=load_button)
    adjust_bttn = bttn_d['adjust'] = tk.Button(bttn_mainfr, textvariable=btxt_adjust, command=adjust_button, state=tk.DISABLED)
    draw_bttn = bttn_d['draw'] = tk.Button(bttn_mainfr, textvariable=btxt_draw, command=draw_button, state=tk.DISABLED)
    save_bttn = bttn_d['save'] = tk.Button(bttn_mainfr, textvariable=btxt_save, command=save_button, state=tk.DISABLED)
    #--widget positioning
    i = 0
    for key, bttn in bttn_d.items():
        bttn.grid(column=i, row=0)
        i += 1
    
    return btn_txt_d, bttn_d

def load_dir_button(*args):
    ''' '''
    
    dirpth = tk.filedialog.askdirectory(initialdir=dname_d['load_dir_pth'].get())
    
    if dirpth == '':
        return
    dirpth += '/'
    dirpth_var = dname_d['load_dir_pth']
    dirpth_var.set(dirpth)
    
    ts_var_pth = dirpth + 'ts_params.txt'
    if os.path.exists(ts_var_pth):
        update_ts_vars(ts_var_pth)
    
    return

def update_ts_vars(pth):
    ''' '''
    
    #load vars
    var_d = orddict()
    with open(pth, mode='r') as file1:
        for line in file1:
            key, val = line.strip().split(';')
            var_d[key] = val
    #udpate dname_d
    for key, val in var_d.items():
        dname_d[key].set(val)
    
    return

def save_dir_button(*args):
    ''' '''
    
    dirpth = tk.filedialog.askdirectory(initialdir=dname_d['save_dir_pth'].get())
    
    if dirpth == '':
        return
    
    dirpth_var = dname_d['save_dir_pth']
    dirpth_var.set(dirpth+'/')
    
    return

def load_button(*args):
    ''' '''
    
    wrap_load_constant_wbs()
    bttn_d['adjust'].config(state=tk.NORMAL) #DISABLED/ACTIVE/NORMAL
    bttn_d['draw'].config(state=tk.NORMAL) #DISABLED/ACTIVE/NORMAL
    bttn_d['save'].config(state=tk.NORMAL) #DISABLED/ACTIVE/NORMAL
    
    return

def adjust_button(*args):
    ''' '''
    
    global adj_nums, adj_loc
    
    #find first case with some data
    adj_nums = find_nums_with_data()
    adj_loc = 0
    #start figure adjust process
    initiate_adjust_window()
    
    return

def initiate_adjust_window():
    ''' '''
    
    global adj_win, adj_mainfr, ctrl_lf1, start_d, adj_num, adj_cbox_var
    
    #create separate window
    try:
        adj_win.destroy()
    except NameError:
        pass
    
    adj_num = adj_nums[adj_loc]
    my_tsd = tsd[adj_num]
    
    adj_win = tk.Toplevel(root)
    adj_win.title('Adjust compress|relax')
    adj_win.geometry('800x600+100+100')
    adj_win.resizable(tk.TRUE,tk.TRUE)
    adj_win.attributes('-topmost', 'true')
    #create mainframe
    adj_mainfr = tk.Frame(adj_win); adj_mainfr.grid()
    #
    ctrl_lf1 = ttk.Labelframe(adj_mainfr, text='controls'); ctrl_lf1.grid(column=0, row=0, sticky=tk.W)
    #add button
    btn = tk.Button(ctrl_lf1, text='Accept -> Next!', command=next_button); btn.grid(column=1, row=0)
    btn = tk.Button(ctrl_lf1, text='Update tsd', command=upd_tsd_button); btn.grid(column=0, row=1)
    #compress and relax start times
    start_d = orddict()
    c_start_var = start_d['c_start'] = tk.DoubleVar(value=my_tsd.var_d['c_start'][1])
    r_start_var = start_d['r_start'] = tk.DoubleVar(value=my_tsd.var_d['r_start'][1])
    tbox_fr = tk.Frame(ctrl_lf1); tbox_fr.grid(column=0, row=0)
    wi = 10
    tbox_widgets(tbox_fr, start_d, wi)
    #combobox for all cases
    adj_cbox_var = tk.IntVar(value=adj_num)
    cbox = ttk.Combobox(ctrl_lf1, textvariable=adj_cbox_var, state='readonly', values=adj_nums, width=4)
    cbox.grid(column=1, row=1)
    cbox.bind('<<ComboboxSelected>>', num_selecter)
    #figure
    adj_fig = update_adjust_figure(my_tsd)

    #-create widgets for figure
    update_adj_fig_widgets(adj_fig)
    
    return

def num_selecter(event):
    ''' '''
    
    global adj_loc, adj_num
    
    adj_num = adj_cbox_var.get()
    adj_loc = adj_nums.index(adj_num)
    print(adj_num)
    adj_upd_starts()
    upd_tsd_button()
    
    return

def update_adj_fig_widgets(adj_fig):
    ''' '''
    
    f_img = tk.Frame(adj_mainfr)
    f_img.grid(column=0, row=2, sticky=(tk.W, tk.N))
    fig_canv = tkagg.FigureCanvasTkAgg(adj_fig, master=f_img)
    fig_c_wid = fig_canv.get_tk_widget()
    #add navigation toolbar
    tbar_fr = tk.Frame(f_img); tbar_fr.grid(column=0, row=0, sticky=tk.W)
    tbar = tkagg.NavigationToolbar2Tk(fig_canv, tbar_fr)
    root.update_idletasks()
    #-set figure size according to window with minimum size of 800x500
    wi = max(adj_win.winfo_width(), 400)
    he = max(adj_win.winfo_height() - ctrl_lf1.winfo_height() - tbar_fr.winfo_height(), 200)
    fig_c_wid.config(width=wi, height=he)
    fig_c_wid.grid(column=0, row=1, sticky=(tk.W, tk.N))
    
    return

def update_adjust_figure(my_tsd):
    ''' '''
    
    #-create figure
    adj_fig = mpl_fig.Figure()
    adj_ax = adj_fig.add_subplot(111)
    adj_ax.set_xlabel('Time(sec)')
    adj_ax.set_ylabel('Stress(MPa)')
    adj_ax.set_xscale('log')
    #-add title
    adj_ax.set_title(adj_num)
    #-add plot data
    datas_pavg = my_tsd.data_d['datas_pavg']
    times = datas_pavg['Time(sec)']
    stresses = datas_pavg['Stress(MPa)']
    #-full line
    line, = adj_ax.plot(times, stresses)
    #-markers
    for key, var in start_d.items():
        val = var.get()
        # print(val)
        val_loc = find_val_loc(times, val)
        s_val = stresses[val_loc]
        adj_ax.plot(val, s_val, color='black', marker='x', markersize=10, markeredgewidth=2)    
    
    return adj_fig

def find_val_loc(vals, findable_val):
    '''tries to find the closest possible location to the findable value. If value not in range of list, will ungracefully raise an error.'''
    
    for i, val in enumerate(vals):
        if val >= findable_val:
            val_loc = i
            break
    
    return val_loc

def upd_tsd_button(*args):
    ''' '''
    
    
    # print('here')
    my_tsd = tsd[adj_num]
    #figure
    adj_fig = update_adjust_figure(my_tsd)

    #-create widgets for figure
    update_adj_fig_widgets(adj_fig)
    
    
    return

def next_button(*args):
    ''' '''
    
    global adj_num
    
    my_tsd = tsd[adj_num]
    values_pavg_d = my_tsd.data_d['datas_pavg']
    times = values_pavg_d['Time(sec)']
    #-update compr and relax starts
    c_s_val = start_d['c_start'].get()
    c_s_loc = find_val_loc(times, c_s_val)
    r_s_val = start_d['r_start'].get()
    r_s_loc = find_val_loc(times, r_s_val)
    c_start = my_tsd.var_d['c_start'] = (c_s_loc, c_s_val)
    r_start = my_tsd.var_d['r_start'] = (r_s_loc, r_s_val)
    #-
    values_c_d, values_r_d = gdh.cut_values_pavg_d(values_pavg_d, c_start, r_start)
    my_tsd.var_d['c_start_stress'] = values_c_d['Stress(MPa)'][0]
    my_tsd.var_d['r_start_stress'] = values_r_d['Stress(MPa)'][0]
    my_tsd.data_d['datas_compress'] = values_c_d
    my_tsd.data_d['datas_relax'] = values_r_d
    #
    adj_open_next_case()
    
    return

def adj_open_next_case():
    ''' '''
    
    global adj_num, adj_loc
    
    adj_loc += 1
    try:
        adj_num = adj_nums[adj_loc]
        adj_upd_starts()
        upd_tsd_button()
    except IndexError:
        print('You found the last accessible figure, feel free to save the data now.')
        adj_win.destroy()
    
    return

def adj_upd_starts():
    ''' '''
    
    my_tsd = tsd[adj_num]
    start_d['c_start'].set(my_tsd.var_d['c_start'][1])
    start_d['r_start'].set(my_tsd.var_d['r_start'][1])
    #
    adj_cbox_var.set(adj_num)
    
    return

def draw_button(*args):
    ''' '''
    
    global draw_loc, draw_nums
    
    #find all cases with some data
    draw_nums = find_nums_with_data()
    draw_loc = 0
    initiate_draw_window()
    
    return

def find_nums_with_data():
    ''' '''
    
    data_nums = []
    for tsd_num, ts_obj in tsd.items():
        try:
            ts_obj.var_d['c_start']
            data_nums.append(tsd_num)
        except KeyError:
            continue
    
    return data_nums

def initiate_draw_window():
    ''' '''
    
    global draw_win, draw_mainfr, draw_ctrl_lf1, draw_d, draw_chk_d, draw_num, draw_fig, draw_ax1, draw_ax2, draw_cbox_var, draw_rb_var, ipol_d
    
    #destroy old window
    try:
        draw_win.destroy()
    except NameError:
        pass
    
    draw_num = draw_nums[draw_loc]
    my_tsd = tsd[draw_num]
    
    draw_win = tk.Toplevel(root)
    draw_win.title('SRX window')
    draw_win.geometry('800x600+100+100')
    draw_win.resizable(tk.TRUE,tk.TRUE)
    # draw_win.attributes('-topmost', 'true')
    #create mainframe
    draw_mainfr = tk.Frame(draw_win); draw_mainfr.grid()
    #
    draw_ctrl_lf1 = ttk.Labelframe(draw_mainfr, text='controls'); draw_ctrl_lf1.grid(column=0, row=0, sticky=tk.W)
    common_fr = tk.Frame(draw_ctrl_lf1); common_fr.grid(column=0, row=0)
    #add button
    btn = tk.Button(common_fr, text='Accept -> Next!', command=draw_next_button)
    btn.grid(column=0, row=0)
    btn = tk.Button(common_fr, text='Skip -> Next!', command=skip_next_button)
    btn.grid(column=0, row=1)
    btn = tk.Button(common_fr, text='Update all', command=draw_upd_button)
    btn.grid(column=2, row=2)
    btn = tk.Button(common_fr, text='Update SRX', command=draw_upd_SRX_button)
    btn.grid(column=2, row=1)
    #straight line control variables
    d1 = orddict()
    d1['a1'] = tk.DoubleVar(value=1.0)
    d1['b1'] = tk.DoubleVar(value=1.0)
    d1['a2'] = tk.DoubleVar(value=1.0)
    d1['b2'] = tk.DoubleVar(value=1.0)
    tbox1_fr = tk.Frame(common_fr); tbox1_fr.grid(column=1, row=0, rowspan=3)
    wi = 8
    tbox_widgets(tbox1_fr, d1, wi)
    #combobox for all cases
    draw_cbox_var = tk.IntVar(value=draw_num)
    cbox = ttk.Combobox(common_fr, textvariable=draw_cbox_var, state='readonly', values=draw_nums, width=4)
    cbox.grid(column=2, row=0)
    cbox.bind('<<ComboboxSelected>>', draw_num_selecter)
    #-checkbox for drawing curves
    chkbox1_fr = ttk.Labelframe(common_fr, text='checks_draw'); chkbox1_fr.grid(column=3, row=0, rowspan=2, sticky=tk.N)
    draw_chk_d = orddict()
    # draw_chk_d['draw_pert'] = tk.IntVar(value=1)
    draw_chk_d['draw_JMAK'] = tk.IntVar(value=1)
    draw_chk_d['draw_Zurob'] = tk.IntVar(value=1)
    cb_widgets(chkbox1_fr, draw_chk_d)
    #-textboxes for x-y -axis limits
    axlim_lf = ttk.Labelframe(common_fr, text='axis limits (empty=AUTO)'); axlim_lf.grid(column=3, row=2)
    d_ax = orddict()
    d_ax['ax1 x'] = tk.StringVar(value='')
    d_ax['ax1 y'] = tk.StringVar(value='')
    d_ax['ax2 x'] = tk.StringVar(value='')
    d_ax['ax2 y'] = tk.StringVar(value='')
    # tbox_ax_fr = tk.Frame(axlim_lf)
    wi = 8
    tbox_widgets(axlim_lf, d_ax, wi, col_len=2)
    
    #Perttula-type fitting widgets
    pert_ctrl_lf = ttk.Labelframe(draw_ctrl_lf1, text='Perttula-JMAK fitting'); pert_ctrl_lf.grid(column=1, row=0)
    #-buttons
    btn = tk.Button(pert_ctrl_lf, text='fit JMAK SRX', command=draw_fit_SRX_button)
    btn.grid(column=1, row=0)
    #-textboxes
    d2 = orddict()
    d2['t50'] = tk.DoubleVar(value=0.0)
    d2['k_JMAK'] = tk.DoubleVar(value=-1e-4)
    d2['n'] = tk.DoubleVar(value=2.0)
    tbox2_fr = tk.Frame(pert_ctrl_lf); tbox2_fr.grid(column=0, row=0, rowspan=3)
    wi = 8
    tbox_widgets(tbox2_fr, d2, wi)
    #-radiobuttons 
    rbox1_fr = tk.Frame(pert_ctrl_lf)
    rbox1_fr.grid(column=3, row=0, sticky=tk.W, rowspan=3)
    rb_name = 'fit options'
    rb_var_names = ['full', 't50_lock', '20-80_SRX']
    draw_rb_var = tk.StringVar(value=rb_var_names[0])
    rb_widgets(rbox1_fr, rb_name, draw_rb_var, rb_var_names)
    #Zurob-type fitting widgets
    Zurob_ctrl_lf = ttk.Labelframe(draw_ctrl_lf1, text='Zurob fitting'); Zurob_ctrl_lf.grid(column=2, row=0)
    #-buttons
    btn = tk.Button(Zurob_ctrl_lf, text='fit Zurob', command=draw_fit_Zurob_button)
    btn.grid(column=1, row=0, sticky=tk.W)
    #-textboxes
    d3 = orddict()
    d3['U_a'] = tk.DoubleVar(value=301e3)
    d3['V_a'] = tk.DoubleVar(value=3.4e-28)
    d3['Q_d'] = tk.DoubleVar(value=68e3)
    d3['k_Z'] = tk.DoubleVar(value=0.05)
    tbox3_fr = tk.Frame(Zurob_ctrl_lf); tbox3_fr.grid(column=0, row=0, rowspan=3)
    wi = 8
    tbox_widgets(tbox3_fr, d3, wi)
    #-interpolate data to ensure constant sampling frequency
    ipol_fr = tk.Frame(Zurob_ctrl_lf); ipol_fr.grid(column=1, row=1)
    ipol_d = orddict()
    var1_name = 'ipol_lim'
    var1 = ipol_d[var1_name] = tk.DoubleVar(value=10.0)
    wi = 5
    tbox_widgets(ipol_fr, ipol_d, wi)
    
    #combine variables
    draw_d = orddict(**d1, **d2, **d3, **d_ax)
    
    #update if already saved variables exist
    wrap_upd_draw_win_vars(my_tsd)
    #figure
    draw_fig, draw_ax1, draw_ax2 = upd_draw_fig1()
    #-widgets for figure
    upd_draw_fig_widgets(draw_fig)
    
    return

def draw_num_selecter(event):
    ''' '''
    
    global draw_num, draw_loc
    
    draw_num = draw_cbox_var.get()
    draw_loc = draw_nums.index(draw_num)
    my_tsd = tsd[draw_num]
    wrap_upd_draw_win_vars(my_tsd)
    draw_upd_button()
    
    return

def upd_draw_fig_widgets(fig1):
    ''' '''
    
    f_img = tk.Frame(draw_mainfr)
    f_img.grid(column=0, row=2, sticky=(tk.W, tk.N))
    fig_canv = tkagg.FigureCanvasTkAgg(fig1, master=f_img)
    fig_c_wid = fig_canv.get_tk_widget()
    #add navigation toolbar
    tbar_fr = tk.Frame(f_img); tbar_fr.grid(column=0, row=0, sticky=tk.W)
    tbar = tkagg.NavigationToolbar2Tk(fig_canv, tbar_fr)
    #-set figure size
    root.update_idletasks()
    wi = max(draw_win.winfo_width(), 400)
    he = max(draw_win.winfo_height() - draw_ctrl_lf1.winfo_height() - tbar_fr.winfo_height(), 200)
    fig_c_wid.config(width=wi, height=he)
    fig_c_wid.grid(column=0, row=1, sticky=(tk.W, tk.N))
    
    return

def upd_draw_fig1():
    ''' '''
    
    #create figure
    fig1 = mpl_fig.Figure()
    ax1 = fig1.add_subplot(211)
    wrap_upd_fig1_ax1(fig1, ax1)
    #second axis
    ax2 = fig1.add_subplot(212)
    ax2_settings(ax2)
    wrap_upd_fig1_ax2(fig1, ax2, 'upd', 'upd')
    fig1.tight_layout()
    
    return fig1, ax1, ax2

def ax1_settings(ax1):
    ''' '''
    
    ax1.set_xlabel('Time(sec)')
    ax1.set_ylabel('Stress(MPa)')
    ax1.set_xscale('log')
    #-add title
    ax1.set_title('%s - relaxation curve' %draw_num)
    #-axis limits
    x_lims = draw_d['ax1 x'].get()
    try:
        x_lims = list(map(float, x_lims.split(',')))
        ax1.set_xlim(x_lims)
    except ValueError:
        pass
    y_lims = draw_d['ax1 y'].get()
    try:
        y_lims = list(map(float, y_lims.split(',')))
        ax1.set_ylim(y_lims)
    except ValueError:
        pass
    
    
    return

def ax2_settings(ax2):
    ''' '''
    
    ax2.set_xlabel('Time(sec)')
    ax2.set_ylabel('Recrystallized portion')
    ax2.set_xscale('log')
    #-add title
    ax2.set_title('%s - SRX curve' %draw_num)    
    #-axis limits
    x_lims = draw_d['ax2 x'].get()
    try:
        x_lims = list(map(float, x_lims.split(',')))
        ax2.set_xlim(x_lims)
    except ValueError:
        pass
    y_lims = draw_d['ax2 y'].get()
    try:
        y_lims = list(map(float, y_lims.split(',')))
        ax2.set_ylim(y_lims)
    except ValueError:
        ax2.set_ylim(0,1)
    
    return

def wrap_upd_fig1_axes(jmak_type, Zurob_type):
    ''' '''
    
    wrap_upd_fig1_ax2(draw_fig, draw_ax2, jmak_type, Zurob_type)
    wrap_upd_fig1_ax1(draw_fig, draw_ax1)
    
    
    return

def wrap_upd_fig1_ax1(fig1, ax1):
    ''' '''
    
    #add plot data
    all_sigs = orddict()
    #-full line
    my_tsd = tsd[draw_num]
    datas_pavg = my_tsd.data_d['datas_relax']
    times = np.array(datas_pavg['Time(sec)'])
    stresses_MPa = datas_pavg['Stress(MPa)']
    all_sigs['exp_rel'] = stresses_MPa
    #-rec line
    # rec_times = np.linspace(0.1, times[-1], num=100)
    a1 = draw_d['a1'].get()
    b1 = draw_d['b1'].get()
    rec_stresses = a1-b1*np.log10(times)
    all_sigs['pert_rec'] = rec_stresses
    #-gg line
    a2 = draw_d['a2'].get()
    b2 = draw_d['b2'].get()
    gg_stresses = a2-b2*np.log10(times)
    all_sigs['pert_gg'] = gg_stresses
    #-Zurob lines
    draw_Zurob_chk = draw_chk_d['draw_Zurob'].get()
    if draw_Zurob_chk:
        fit_vars = Zurob_get_fit_vars()
        a2 = draw_d['a2'].get()
        b2 = draw_d['b2'].get()
        gg_vars = [a2*1e6, b2*1e6]
        stresses_Pa = np.array(stresses_MPa)*1e6
        sig_start = stresses_Pa[0]
        T_degC = my_tsd.var_d['nom_T']
        sig_y = find_sig_y(times, stresses_Pa, gg_stresses*1e6)
        if sig_y != '':
            try:
                sig_ggs = a2-b2*np.log10(times_Z)
                mat_type = dname_d['mat_type'].get()
                vals_d, sig_recs, sig_rels, sig_ggs, SRXs = gdh.wrap_Zurob_runner(times_Z, sig_ggs*1e6, locs_Z, fit_vars, gg_vars, sig_start, T_degC, sig_y, mat_type)
                all_sigs['Zurob_rec'] = np.array(sig_recs)*1e-6
                all_sigs['Zurob_rel'] = np.array(sig_rels)*1e-6
            except NameError:
                pass

    upd_draw_fig_ax1(fig1, ax1, times, all_sigs)

    
    return

def upd_draw_fig_ax1(fig1, ax1, times, all_sigs):
    ''' '''
    
    #reset axis
    ax1.clear()
    ax1_settings(ax1)
    #add plot data
    for lbl, sigs in all_sigs.items():
        line, = ax1.plot(times, sigs, label=lbl)
    #
    ax1.legend()
    fig1.canvas.draw()    
    
    return

def upd_draw_fig1_ax2(fig1, ax2, times, all_SRXs):
    ''' '''
    
    ax2.clear()
    ax2_settings(ax2)
    
    #add plot data
    for lbl, SRXs in all_SRXs.items():
        line, = ax2.plot(times, SRXs, label=lbl)
    #add t50 marker
    for lbl, t50_val in t50_d.items():
        ax2.plot(t50_val, 0.5, color='black', marker='x', markersize=10, markeredgewidth=3)#, label=lbl)
    # ax2.plot(draw_d['t50'].get(), 0.5, color='black', marker='x', markersize=10, markeredgewidth=3, label='t50')
    ax2.legend()
    fig1.canvas.draw()
    
    return

def wrap_upd_fig1_ax2(fig1, ax2, jmak_type, Zurob_type):
    ''' '''
    
    global t50_d
    
    my_tsd = tsd[draw_num]
    datas_pavg = my_tsd.data_d['datas_relax']
    times = datas_pavg['Time(sec)']
    stresses = datas_pavg['Stress(MPa)']
    #-draw checks
    # draw_pert_chk = draw_chk_d['draw_pert'].get()
    draw_jmak_chk = draw_chk_d['draw_JMAK'].get()
    draw_Zurob_chk = draw_chk_d['draw_Zurob'].get()
    draw_SRXs = orddict()
    #-perttula SRX
    np_times = np.array(times)
    np_stresses_MPa = np.array(stresses)
    a1 = draw_d['a1'].get()
    b1 = draw_d['b1'].get()
    a2 = draw_d['a2'].get()
    b2 = draw_d['b2'].get()
    rec_stresses = a1-b1*np.log10(np_times)
    gg_stresses = a2-b2*np.log10(np_times)
    pert_SRXs = gdh.pert_SRX_basic_eq(np_times, np_stresses_MPa, rec_stresses, gg_stresses)
    draw_SRXs['pert'] = pert_SRXs
    t50_val = find_t50_val(np_times, pert_SRXs)
    #get jmak curve
    if draw_jmak_chk:
        jmak_SRXs = wrap_jmak_curve(jmak_type, np_times, pert_SRXs, t50_val)
        draw_SRXs['JMAK'] = jmak_SRXs
    #get Zurob curve
    if draw_Zurob_chk:
        sig_recs, Zurob_SRXs = wrap_Zurob_curve(Zurob_type)
        if len(Zurob_SRXs) > 0:
            draw_SRXs['Zurob'] = Zurob_SRXs
            Zpert_SRXs = gdh.pert_SRX_basic_eq(np_times, np_stresses_MPa, np.array(sig_recs)*1e-6, gg_stresses)
            draw_SRXs['Zpert'] = Zpert_SRXs
    #find and set t50
    t50_d = wrap_t50_finder(np_times, draw_SRXs)
    draw_d['t50'].set(t50_d['t50_pert'])
    #update fig1 ax2
    upd_draw_fig1_ax2(fig1, ax2, np_times, draw_SRXs)
    
    return

def wrap_t50_finder(np_times, draw_SRXs):
    ''' '''
    
    t50_d = orddict()
    for lbl, SRXs in draw_SRXs.items():
        t50_val = find_t50_val(np_times, SRXs)
        t50_d['t50_%s' %lbl] = t50_val
    
    return t50_d

def find_t50_val(times, SRXs):
    ''' '''
    
    #find locations where SRX~=0.5
    t50s = []
    for i, val in enumerate(SRXs):
        t50_chk = val > 0.49 and val < 0.51
        if t50_chk:
            time = times[i]
            t50s.append(time)
    len_chk = len(t50s) >= 2
    #calculate t50
    if len_chk: #many datapoints around t50, averaging seems to work nicely
        t50_val = sum(t50s) / len(t50s)
    else: #0-1 datapoints near t50, linear interpolation used
        t50_val = wrap_lin_interpol(times, SRXs, 0.5)
    
    return t50_val

def wrap_lin_interpol(xs, ys, y_get):
    ''' '''
    
    y_range_chk = y_get > min(ys) and y_get < max(ys)
    if not y_range_chk:
        return ''
    for i, y in enumerate(ys):
        y_chk = y >= y_get
        if y_chk:
            y1 = ys[i-1]
            y2 = y
            x1 = xs[i-1]
            x2 = xs[i]
            break
    
    x_get = (y_get - y1) / ((y2-y1)/(x2-x1)) + x1
    
    return x_get

def find_sig_y(np_times, stresses_exp, gg_stresses):
    ''' '''
    
    sig_y = ''
    for i, sig_gg in enumerate(gg_stresses):
        sig_exp = stresses_exp[i]
        time = np_times[i]
        sig_chk = sig_gg > sig_exp and time > 0.001
        if sig_chk:
            sig_y = sig_gg
            break
    
    return sig_y

def wrap_Zurob_curve(Zurob_type):
    ''' '''
    
    a2 = draw_d['a2'].get()
    b2 = draw_d['b2'].get()
    gg_vars = [a2*1e6, b2*1e6]
    try:
        sig_start = stresses_Z[0]
    except NameError:
        return [], []
    T_degC = tsd[draw_num].var_d['nom_T']
    gg_stresses = gg_vars[0]-gg_vars[1]*np.log10(times_Z)
    sig_y = find_sig_y(times_Z, stresses_Z, gg_stresses)
    mat_type = dname_d['mat_type'].get()
    if sig_y == '':
        return [], []
    if Zurob_type == 'fit':
        print('-#-#-\nZurob fitting')
        fit_vars_0 = [301E3,3.4375E-28,68000,0.05]
        fit_vars = gdh.wrap_Zurob_fitter(T_degC, fit_vars_0, gg_vars, sig_start, sig_y, times_Z, stresses_Z, gg_stresses, mat_type)
        Zurob_set_fit_vars(fit_vars)
        print('Fitting complete!\n#-#-#')
    elif Zurob_type == 'upd':
        # fit_vars = [3.154016417091605e+05, 9.296916030166785e-28, 1.066872808468765e+05, 0.050752308144109]
        fit_vars = Zurob_get_fit_vars()
    
    vals_d, sig_recs, sig_rels, sig_ggs, SRXs = gdh.wrap_Zurob_runner(times_Z, gg_stresses, locs_Z, fit_vars, gg_vars, sig_start, T_degC, sig_y, mat_type)
    
    return sig_recs, SRXs

def Zurob_set_fit_vars(fit_vars):
    ''' '''
    
    draw_d['U_a'].set(fit_vars[0])
    draw_d['V_a'].set(fit_vars[1])
    draw_d['Q_d'].set(fit_vars[2])
    draw_d['k_Z'].set(fit_vars[3])
    
    return

def Zurob_get_fit_vars():
    ''' '''
    
    U_a = draw_d['U_a'].get()
    V_a = draw_d['V_a'].get()
    Q_d = draw_d['Q_d'].get()
    k_Z = draw_d['k_Z'].get()
    fit_vars = [U_a, V_a, Q_d, k_Z]
    
    return fit_vars

def wrap_jmak_curve(jmak_type, np_times, pert_SRXs, t50_val):
    ''' '''
    
    if jmak_type == 'fit':
        fit_type = draw_rb_var.get()
        print('-#-#-\nJMAK fitting with option: %s' %fit_type)
        if fit_type == 'full':
            #-fitted jmak
            k0 = draw_d['k_JMAK'].get()
            n0 = draw_d['n'].get()
            fit_vals = gdh.wrap_jmak_fitter(np_times, pert_SRXs, [k0, n0], fit_type, t50_val)
            k, n = fit_vals
            draw_d['k_JMAK'].set(k)
            draw_d['n'].set(n)
            jmak_SRXs = 1-np.exp(k*np_times**n)
        elif fit_type == 't50_lock':
            n0 = draw_d['n'].get()
            fit_vals = gdh.wrap_jmak_fitter(np_times, pert_SRXs, [n0], fit_type, t50_val)
            n, = fit_vals
            k = np.log(0.5) / (t50_val**n)
            draw_d['k_JMAK'].set(k)
            draw_d['n'].set(n)
            jmak_SRXs = 1-np.exp(np.log(0.5)*(np_times/t50_val)**n)
        elif fit_type == '20-80_SRX':
            #-fitted jmak
            k0 = draw_d['k_JMAK'].get()
            n0 = draw_d['n'].get()
            fit_vals = gdh.wrap_jmak_fitter(np_times, pert_SRXs, [k0, n0], fit_type, t50_val)
            k, n = fit_vals
            draw_d['k_JMAK'].set(k)
            draw_d['n'].set(n)
            jmak_SRXs = 1-np.exp(k*np_times**n)
        print('fitting complete!\n#-#-#')
    elif jmak_type == 'upd':
        k = draw_d['k_JMAK'].get()
        n = draw_d['n'].get()
        jmak_SRXs = 1-np.exp(k*np_times**n)
    
    return jmak_SRXs

def draw_fit_SRX_button(*args):
    ''' '''
    
    jmak_type = 'fit'
    Zurob_type = 'upd'
    wrap_upd_fig1_axes(jmak_type, Zurob_type)
    
    return

def draw_fit_Zurob_button(*args):
    ''' '''
    
    jmak_type = 'upd'
    Zurob_type = 'fit'
    wrap_upd_fig1_axes(jmak_type, Zurob_type)
    
    return

def draw_upd_SRX_button(*args):
    ''' '''
    
    jmak_type = 'upd'
    Zurob_type = 'upd'
    wrap_upd_fig1_axes(jmak_type, Zurob_type)
    
    return

def draw_open_next_case():
    ''' '''
    
    global draw_num, draw_loc
    
    #open next case
    draw_loc += 1
    try:
        draw_num = draw_nums[draw_loc]
        my_tsd = tsd[draw_num]
        wrap_upd_draw_win_vars(my_tsd)
        draw_upd_button()
    except IndexError:
        print('Last figure processed, Perttula-type line drawing completed!')
        draw_win.destroy()    
    
    return

def wrap_SRX_pert_jmak_eqs(my_tsd):
    ''' '''
    
    d1 = my_tsd.data_d
    d2 = my_tsd.SRX_d
    d3 = my_tsd.SRX_data_d
    
    #save all SRX vars
    for key, val in draw_d.items():
        if key in ('t50', 'ax1 x', 'ax1 y', 'ax2 x', 'ax2 y'):
            continue
        d2[key] = val.get()
    for key, val in t50_d.items():
        d2[key] = val
    
    datas = d1['datas_relax']
    times = np.array(datas['Time(sec)'])
    #stresses
    stresses = np.array(datas['Stress(MPa)'])
    a1, b1, a2, b2 = d2['a1'], d2['b1'], d2['a2'], d2['b2']
    rec_stresses = a1-b1*np.log10(times)
    gg_stresses = a2-b2*np.log10(times)
    #SRX
    #-perttula-type
    pert_SRXs = gdh.pert_SRX_basic_eq(times, stresses, rec_stresses, gg_stresses)
    k, n = d2['k_JMAK'], d2['n']
    #-jmak
    jmak_SRXs = 1-np.exp(k*times**n)
    #-Zurob
    fit_vars = [d2['U_a'], d2['V_a'], d2['Q_d'], d2['k_Z']]
    gg_vars = [a2*1e6, b2*1e6]
    sig_start = stresses[0]*1e6
    T_degC = my_tsd.var_d['nom_T']
    sig_y = find_sig_y(times, stresses*1e6, gg_stresses*1e6)
    if sig_y != '':
        sig_ggs = a2-b2*np.log10(times_Z)
        mat_type = dname_d['mat_type'].get()
        vals_d, sig_recs, sig_rels, sig_ggs, Zurob_SRXs = gdh.wrap_Zurob_runner(times_Z, sig_ggs*1e6, locs_Z, fit_vars, gg_vars, sig_start, T_degC, sig_y, mat_type)
        Zpert_SRXs = gdh.pert_SRX_basic_eq(times, stresses, np.array(sig_recs)*1e-6, gg_stresses)
    else:
        sig_recs, sig_rels, Zurob_SRXs, Zpert_SRXs = [], [], [], []
    #
    d3['Time'] = times
    d3['Stress'] = stresses
    d3['rec_stress'] = rec_stresses
    d3['gg_stress'] = gg_stresses
    d3['Zurob_rec_stress'] = np.array(sig_recs)*1e-6
    d3['Zurob_rel_stress'] = np.array(sig_rels)*1e-6
    d3['pert_SRX'] = pert_SRXs
    d3['jmak_SRX'] = jmak_SRXs
    d3['Zurob_SRX'] = Zurob_SRXs
    d3['Zpert_SRX'] = Zpert_SRXs
    
    return 

def draw_next_button(*args):
    ''' '''
    
    
    my_tsd = tsd[draw_num]
    #save SRX data
    wrap_SRX_pert_jmak_eqs(my_tsd)
    print('Added %d' %draw_num)
    #next case
    draw_open_next_case()
    
    return

def skip_next_button(*args):
    ''' '''
    
    print('Skipped %d' %draw_num)
    
    draw_open_next_case()
    
    return

def draw_upd_button(*args):
    ''' '''
    
    global draw_fig, draw_ax1, draw_ax2, times_Z, stresses_Z, locs_Z
    
    my_tsd = tsd[draw_num]
    times_Z, stresses_Z, locs_Z = wrap_ipolate_data_Z(my_tsd)
    draw_fig, draw_ax1, draw_ax2 = upd_draw_fig1()
    upd_draw_fig_widgets(draw_fig)
    draw_cbox_var.set(draw_num)
    
    return

def wrap_upd_draw_win_vars(my_tsd):
    ''' '''
    
    d1 = my_tsd.SRX_d
    #straight line variables
    try:
        a1, b1, a2, b2 = d1['a1'], d1['b1'], d1['a2'], d1['b2']
    except KeyError:
        a1, b1, a2, b2 = 1.0, 1.0, 1.0, 1.0
    draw_d['a1'].set(a1)
    draw_d['a2'].set(a2)
    draw_d['b1'].set(b1)
    draw_d['b2'].set(b2)
    #JMAK variables
    try:
        t50, k_JMAK, n = d1['t50'], d1['k_JMAK'], d1['n']
    except KeyError:
        t50, k_JMAK, n = 0.0, -1e-4, 2.0
    draw_d['t50'].set(t50)
    draw_d['k_JMAK'].set(k_JMAK)
    draw_d['n'].set(n)
    #Zurob variables
    try:
        U_a, V_a, Q_d, k_Z = d1['U_a'], d1['V_a'], d1['Q_d'], d1['k_Z']
    except KeyError:
        U_a, V_a, Q_d, k_Z = 301e3, 3.4e-28, 68e3, 0.05
    draw_d['U_a'].set(U_a)
    draw_d['V_a'].set(V_a)
    draw_d['Q_d'].set(Q_d)
    draw_d['k_Z'].set(k_Z)
    
    return

def wrap_ipolate_data_Z(my_tsd):
    ''' '''
    
    datas_relax = my_tsd.data_d['datas_relax']
    times = datas_relax['Time(sec)']
    stresses = datas_relax['Stress(MPa)']
    t_diffs, s_freqs = get_s_freq(times)
    # ipol_chk = ipol_d['ipol_chk'].get()
    ipol_lim = ipol_d['ipol_lim'].get()
    times_Z, stresses_Z, locs_Z = ipol_data_Z(times, stresses, s_freqs, ipol_lim)
    # if ipol_chk:
        # times_Z, stresses_Z, locs_Z = ipol_data_Z(times, stresses, s_freqs, ipol_lim)
    # else:
        # times_Z, stresses_Z, locs_Z = np.array(times), np.array(stresses)*1e6, ''
    
    return times_Z, stresses_Z, locs_Z

def ipol_data_Z(times, stresses, s_freqs, ipol_lim):
    ''' '''
    
    times_Z, stresses_Z = [times[0]], [stresses[0]]
    prev_time, prev_stress = times[0], stresses[0]
    locs_Z = [0]
    init_s_freq = s_freqs[1]
    num = 0
    for i, time in enumerate(times):
        stress = stresses[i]
        s_freq = s_freqs[i]
        try:
            chk1 = init_s_freq > s_freq
        except TypeError:
            continue
        chk2 = time < ipol_lim
        chk3 = s_freq > 0
        if chk1 and chk2 and chk3:
            ipol_range = int(init_s_freq // s_freq)
            num += ipol_range - 1
            range_ipolator(times_Z, prev_time, time, ipol_range)
            range_ipolator(stresses_Z, prev_stress, stress, ipol_range)
            
        
        times_Z.append(time)
        stresses_Z.append(stress)
        prev_time = time
        prev_stress = stress
        num += 1
        locs_Z.append(num)
    
    times_Z = np.array(times_Z)
    stresses_Z = np.array(stresses_Z)*1e6
    
    return times_Z, stresses_Z, locs_Z

def range_ipolator(vals, prev_val, val, ipol_range):
    ''' '''
    
    for j in range(1, ipol_range):
        ipol_val = prev_val + (j/ipol_range * (val-prev_val))
        # print('prev_val', prev_val)
        # print('val', val)
        # print('ipol_val', ipol_val)
        # input()
        vals.append(ipol_val)
    
    return

def get_s_freq(times):
    ''' '''
    
    t_diffs = ['']
    s_freqs = ['']
    prev_time = times[0]
    for time in times[1:]:
        t_diff = time - prev_time
        t_diffs.append(t_diff)
        try:
            s_freqs.append(1/t_diff)
        except ZeroDivisionError:
            s_freqs.append('')
        prev_time = time
    
    return t_diffs, s_freqs

def save_button(*args):
    ''' '''
    
    wrap_save_constant_wbs()
    
    return


if __name__ == '__main__':
    
    # test_mode = 'single_wb' #single_wb|no
    
    root_path = ''
    if not os.path.exists(root_path):
        root_path = ''
    
    my_plat = sys.platform
    if my_plat == 'win32': #windows users
        ts_dir = ''
        ts_path = root_path + ts_dir
        #load_dir_pth_init = ts_path + 'tulokset/00_kaikki_csv/'
        load_dir_pth_init = ts_path + 'data/test_csv/'
        save_dir_pth_init = ts_path + 'data/results_handling/'
    else: #mac users
        load_dir_pth_init = ''
        save_dir_pth_init = ''
    #old raex400 - relaxation-1
    temps = '[850, 900, 950, 1000, 1050]'
    strains = '[0.2, 0.3, 0.4]'
    SRs = '[1, 10, 40]'
    hold_times = '[200]'
    #Steel 1523 with Mohammed
    temps = '[850, 950, 1050, 1150, 1250]'
    strains = '[0.2, 0.4, 0.6]'
    SRs = '[10]'
    hold_times = '[0,30,60]'
            
    
    ##create a main window and a title for it
    root = tk.Tk()
    root.title("Relax reader!")
    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)
    #-define min size for main window
    root.minsize(1000,500)
    #-create a mainframe inside the main window
    mainfr = tk.Frame(root)
    mainfr.grid(column=0, row=0, sticky=(tk.N,tk.W,tk.E,tk.S))
    
    #-labelframe for controls
    ctrl_lf  = ttk.Labelframe(mainfr, text='Controls')
    ctrl_lf.grid(column=0, row=0, sticky=(tk.W,tk.E, tk.N))
    #-frame for canvases
    canv_fr = tk.Frame(mainfr)
    canv_fr.grid(column=0, row=1, sticky=(tk.W,tk.E))
    
    #create buttons
    btn_txt_d, bttn_d = create_buttons(ctrl_lf)
    #create variable holders
    dname_d = create_textboxes(ctrl_lf)
    
    
    #opens the main window for the user
    root.mainloop()

    
    
    
    
    
    
    
    
    
