'''
All kinds of functions for excel usage.

'change required for openpyxl module code to change chart area properties
Python27\Lib\site-packages\openpyxl\chart\_chart.py
    #cs = ChartSpace(chart=container)
    cs = ChartSpace(chart=container, spPr=self.graphical_properties)
'

'''

from collections import OrderedDict as orddict
import xlsxwriter as xw
import os
import openpyxl as opyx
import numpy as np
import math

############### openpyxl functions ##################################

def opyx_set_col_w(ws):
    '''automatically set a somewhat nice estimation of the required column width.'''
    
    dims = {}
    all_rows = ws.rows
    for row in all_rows:
        for cell in row:
            if cell.value:
                try:
                    dims[cell.column_letter] = max((4, dims[cell.column_letter], len(str(cell.value))+2))
                except KeyError:
                    dims[cell.column_letter] = max((4, len(str(cell.value))+2))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value
    
    return

## create workbook and save data

def opyx_dict_to_wb(wb_pth, wb_dict, num_dec='', auto_col_w='yes'):
    '''creates an excel workbook and adds the contents of a dict with syntax {ws_name:{header:val_list}}.'''
    
    #creates workbook object
    wb = opyx.Workbook()
    #removes default sheet
    wb.remove(wb['Sheet'])
    for ws_name, ws_d in wb_dict.items():
        ws = wb.create_sheet(str(ws_name))
        opyx_dict_to_ws(ws, ws_d, num_dec)
        if auto_col_w == 'yes':
            opyx_set_col_w(ws)

    wb.save(wb_pth)
    
    return wb

def opyx_dh_dict_to_wb(wb_pth, wb_dict, num_dec='', auto_col_w='yes', single_ws_nums=[]):
    ''' '''
    
    #creates workbook object
    wb = opyx.Workbook()
    #removes default sheet
    wb.remove(wb['Sheet'])
    ws_num = 0
    for ws_name, ws_d in wb_dict.items():
        ws = wb.create_sheet(str(ws_name))
        if ws_num in single_ws_nums:
            opyx_dict_to_ws(ws, ws_d, num_dec)
        else:
            opyx_dict_doubleheader_to_ws(ws, ws_d, num_dec)
        if auto_col_w == 'yes':
            opyx_set_col_w(ws)
        ws_num += 1

    wb.save(wb_pth)
    
    return wb

def opyx_dict_to_ws(ws, ws_dict, num_dec):
    '''Adds the contents of a dictionary to a worksheet. dict needs to have syntax {header: value_list}.'''
    
    col_val = 1
    for this_header, these_vals in ws_dict.items():
        # print(this_header)
        # try:
            # akof.dict_printer(these_vals)
            # raise myerror
        # except AttributeError:
            # pass
        opyx_excel_data_writer(ws, str(this_header), these_vals, col_val, num_dec)
        col_val += 1
    
    return

def opyx_dict_doubleheader_to_ws(ws, ws_d, num_dec):
    ''' '''
    
    col_val = 1
    for doubleheader, sub_d in ws_d.items():
        dhead_cell = ws.cell(row=1, column=col_val)
        dhead_cell.value = str(doubleheader)
        old_col_val = col_val
        for head, val_list in sub_d.items():
            opyx_excel_data_writer(ws, str(head), val_list, col_val, num_dec, row_start=2)
            col_val +=1
        ws.merge_cells(start_row=1, end_row=1, start_column=old_col_val, end_column=col_val-1)
        
        
    
    return

def opyx_excel_data_writer(ws, header, vals, col_val, num_dec, row_start=1):
    '''writes a header and data to a column.'''
    
    header_cell = ws.cell(row=row_start, column=col_val)
    header_cell.value = header
    vals_len = len(vals)
    for i in range(row_start+1, vals_len+row_start+1):
        this_cell = ws.cell(row=i, column=col_val)
        this_val = vals[i-(row_start+1)]
        this_cell.value = this_val
        # try:
            # this_cell.value = this_val
        # except ValueError:
            # print(this_val)
            # quit()
            # this_cell.value = float(this_val)
            
        if num_dec != '':
            if type(this_val) == type(np.float64(0.0)) or type(this_val) == type(0.0):
                zero_caller = '%.'+str(num_dec)+'f'
                this_cell.number_format = zero_caller %0
    
    return

def wrap_all_data_to_wb(wb_pth, tsd, data_str, var_lbls, log10_chk='no', cr_starts='no'):
    ''' '''
    
    print('saving all data to single workbook...')
    #-combine variables
    all_vars_d = orddict()
    for num, ts_obj in tsd.items():
        for var_lbl in var_lbls:
            if var_lbl == 'num':
                var_val = num
            elif var_lbl in ('c_start', 'r_start'):
                try:
                    var_val = ts_obj.var_d[var_lbl][1]
                except KeyError:
                    var_val = ''
            else:
                try:
                    var_val = ts_obj.var_d[var_lbl]
                except KeyError:
                    var_val = ''
            try:
                all_vars_d[var_lbl].append(var_val)
            except KeyError:
                all_vars_d[var_lbl] = [var_val]
    #-create wb dict
    wb_d = orddict()
    wb_d['variables'] = all_vars_d
    for num, ts_obj in tsd.items():
        try:
            wb_d['%d'%num] = ts_obj.data_d[data_str]
        except KeyError:
            continue
    
    #create workbook
    wb = opyx_dict_to_wb(wb_pth, wb_d)
    print('Workbook completed, creating charts...')
    #create charts
    for sheetname in wb.sheetnames:
        #skip variable ws
        if sheetname == 'variables':
            continue
        #all samex charts
        ws = wb[sheetname]
        charts = samex_data_to_little_charts(ws, log10_chk)
        #stress-strain chart
        x_lbl = 'Strain'
        y_lbl = 'Stress(MPa)'
        xy_to_little_chart(ws, charts, x_lbl, y_lbl)
        #SR-strain chart
        y_lbl = 'Strain rate(1/s)'
        xy_to_little_chart(ws, charts, x_lbl, y_lbl)
        #update stress chart to include compress and relax starting points
        if 'c_start' in var_lbls:
            cr_points_to_chart(charts, ws, wb['variables'])
        
    #save charts
    wb.save(wb_pth)
    print('%s Finished!' %wb_pth)
    
    return

##create charts into workbook

def charts_to_const_wb(wb, x_var):
    '''Adds charts to a workbook with constant values.'''
    
    for ws in wb:
    
        max_row = ws.max_row
        max_col = ws.max_column
        c1 = opyx.chart.ScatterChart()
        c1.style = 13
        c1.width = 45
        c1.height = 18
        c1.y_axis.title = 'Stress [MPa]'
        c1.x_axis.title = x_var
        if x_var == 'Time':
            c1.x_axis.scaling.logBase = 10
            x_minor_ticks_ = 'yes'
        else:
            x_minor_ticks_ = 'no'
        chart_default_settings(c1, x_minor_ticks=x_minor_ticks_)
        
        start_col = 1
        skipper = 1
        end_col = max_col // 2 
        opyx_dhdata2chart_new(ws, c1, start_col, skipper, end_col, max_row)
        # for i in range(1, range_max):
            # x_col, y_col = (i * 2 - 1), (i * 2)
            
            # x_vals = opyx.chart.Reference(ws, min_col=x_col, min_row=2, max_row=max_row)
            # y_vals = opyx.chart.Reference(ws, min_col=y_col, min_row=1, max_row=max_row)
            # this_series = opyx.chart.Series(y_vals, x_vals, title_from_data=True)
            # c1.series.append(this_series)
            
        ws.add_chart(c1, 'B4')

    return

def set_axis_settings(ax_obj, title, ax_lims):
    '''Settings for an axis, title and upper and lower limits.'''
    
    if title != 0:
        ax_obj.title = title
    if ax_lims != 0:
        ax_obj.scaling.min = ax_lims[0]
        ax_obj.scaling.max = ax_lims[1]
    
    return

def series_line2markers(ser1, m_size=10.0, m_w_pt=2, marker_color='default'):
    ''' '''
    
    # marker_size = 20.0
    # marker_width = 50000
    #symbols -> {dot, plus, triangle, x, picture, star, diamond, square, circle, dash, auto
    ser1.marker=opyx.chart.marker.Marker(symbol='x', size=m_size)
    ser1.graphicalProperties.line.noFill=True
    m_w_emu = pt2emu(m_w_pt)
    ser1.marker.graphicalProperties.line.width = m_w_emu # width in EMUs
    if marker_color == 'black':
        ser1.marker.graphicalProperties.line.solidFill = opyx.drawing.colors.ColorChoice(prstClr='black')
    
    return

def pt2emu(w_pt):
    ''' '''
    
    w_emu_1p = int(10000/0.79) #emu size to get 1p width
    w_emu = int(w_emu_1p * w_pt)
    
    return w_emu

def opyx_dhdata2chart_new(ws, c1, start_col, skipper, end_col, max_row, xy_locs=(1,2), stack_len=2):
    ''' '''
    
    x_loc, y_loc = xy_locs
    serieses = []
    for i in range(start_col, end_col+1, skipper):
        # print('%d, ' %i,)
        x_col, y_col = col_calc(i, x_loc, stack_len), col_calc(i, y_loc, stack_len)
        x_vals = opyx.chart.Reference(ws, min_col=x_col, min_row=3, max_row=max_row)
        y_vals = opyx.chart.Reference(ws, min_col=y_col, min_row=3, max_row=max_row)
        # s_title = ws.cell(row=1, column=x_col).value
        s_title = opyx.chart.Reference(ws, min_col=x_col, min_row=1)
        my_s = opyx.chart.Series(y_vals, x_vals, title=s_title)
        # series_line2markers(my_s)
        #HOX! by default, title is saved as a string, so even a cell refence is converted into string.
        my_s.title.v = None #remove string to prevent using it as series title
        my_s.title.strRef = opyx.chart.data_source.StrRef(s_title) #add cell reference as series title
        c1.series.append(my_s)
        serieses.append(my_s)
    
    return serieses

def col_calc(i, loc, stack_len):
    '''little function that calculates a column-wise location for data.
        Returns said location.'''
    
    this_col = i * stack_len - stack_len + loc
    
    return this_col

def cell_loc_gen(col, row):
    '''generate excel worksheet locations from numerical locations. 1-based. includes two levels of column names
        return anch'''
    
    alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    len_a = len(alphabet)
    col_s = ''
    if col >= len_a:
        col_s += alphabet[col//len_a-1]
    col_s += alphabet[col%len_a]
    anch = '%s%d' %(col_s, row)
    
    return anch

def samex_data_to_little_charts(ws, log10_chk='no', axes_set_lb_1='yes', chart_type='scatter', nums_in_row=3):
    ''' '''
    
    max_row = ws.max_row
    max_col = ws.max_column
    #
    x_vals = opyx.chart.Reference(ws, min_col=1, min_row=2, max_row=max_row)
    charts = []
    for i in range(max_col-1):
        y_col = i+2
        #create chart
        if chart_type == 'scatter':
            c1 = opyx.chart.ScatterChart()
        elif chart_type == 'line':
            c1 = opyx.chart.LineChart()
        elif chart_type == 'bar':
            c1 = opyx.chart.BarChart()
        c1.style = 2
        c1.width = 15.0
        c1.height = 7.8
        chart_default_settings(c1, axes_set_lb=axes_set_lb_1)
        c1.x_axis.title = ws.cell(row=1, column=1).value #'Thickness [m]'
        c1.y_axis.title = ws.cell(row=1, column=y_col).value #var
        if log10_chk == 'yes':
            c1.x_axis.scaling.logBase = 10
        #remove legend
        c1.legend = None
        #add series
        y_vals = opyx.chart.Reference(ws, min_col=y_col, min_row=1, max_row=max_row)
        if chart_type == 'scatter':
            my_series = opyx.chart.Series(y_vals, x_vals, title_from_data=True)
            c1.series.append(my_series)
        elif chart_type in ('line', 'bar'):
            c1.add_data(y_vals, titles_from_data=True)
            c1.set_categories(x_vals)
        #add chart to worksheet
        anch = cell_loc_gen((i%nums_in_row)*9+max_col+1,(i//nums_in_row)*15+1)
        ws.add_chart(c1, anchor=anch)
        charts.append(c1)
    
    return charts

def xy_to_little_chart(ws, charts, x_lbl, y_lbl):
    ''' '''
    
    max_row = ws.max_row
    max_col = ws.max_column
    #create chart
    c1 = opyx.chart.ScatterChart()
    c1.style = 2
    c1.width = 15.0
    c1.height = 7.8
    chart_default_settings(c1)
    c1.x_axis.title = x_lbl
    c1.y_axis.title = y_lbl
    # if log10_chk == 'yes':
        # c1.x_axis.scaling.logBase = 10
    #remove legend
    c1.legend = None
    #add series
    x_col = find_label_loc(ws, x_lbl)
    y_col = find_label_loc(ws, y_lbl)
    x_vals = opyx.chart.Reference(ws, min_col=x_col, min_row=2, max_row=max_row)
    y_vals = opyx.chart.Reference(ws, min_col=y_col, min_row=1, max_row=max_row)
    my_series = opyx.chart.Series(y_vals, x_vals, title_from_data=True)
    c1.series.append(my_series)
    # if y_lbl == 'Strain rate(1/s)':
        # ser1 = opyx.chart.series_factory.SeriesFactory([1.0,1.0],[0.0, 0.4])
        # c1.series.append(my_series)
    #add chart to worksheet
    # i = len(list(ws.values)[0]) - 1
    i = len(charts)
    anch = cell_loc_gen((i%3)*9+max_col+1,(i//3)*15+1)
    ws.add_chart(c1, anchor=anch)
    charts.append(c1)
    
    return

def SRX_data_to_little_charts(ws, log10_chk='yes'):
    ''' '''
    
    max_row = ws.max_row
    max_col = ws.max_column
    for i, type in enumerate(['stress', 'SRX']):
        #create chart
        c1 = opyx.chart.ScatterChart()
        c1.style = 2
        c1.width = 20.0
        c1.height = 7.8
        chart_default_settings(c1)
        c1.x_axis.title = 'Time [log s]'
        if log10_chk == 'yes':
            c1.x_axis.scaling.logBase = 10
        #-serieses
        x_vals = opyx.chart.Reference(ws, min_col=1, min_row=2, max_row=max_row)
        if type == 'stress':
            y_cols = [2,3,4,5,6]
            c1.y_axis.title = 'Stress [MPa]'
        elif type == 'SRX':
            y_cols = [7,8,9,10]
            c1.y_axis.title = 'SRX progression'
        for y_col in y_cols:
            #add series
            y_vals = opyx.chart.Reference(ws, min_col=y_col, min_row=1, max_row=max_row)
            my_series = opyx.chart.Series(y_vals, x_vals, title_from_data=True)
            c1.series.append(my_series)
        #add chart to worksheet
        # anch = cell_loc_gen((i%3)*9+max_col+1,(i//3)*15+1)
        anch = cell_loc_gen(max_col+1,i*15+1)
        ws.add_chart(c1, anchor=anch)    
        
    return

def find_label_loc(ws, lbl_name):
    ''' '''
    
    lbls = list(ws.values)[0]
    lbl_loc = lbls.index(lbl_name) + 1
    
    return lbl_loc

def cr_points_to_chart(charts, ws, vars_ws):
    ''' '''
    
    #find correct chart
    lbls = list(ws.values)[0]
    c_loc = lbls.index('Stress(MPa)')-1
    c1 = charts[c_loc]
    #create markers
    ws_num = int(ws.title)
    #-compression start
    vars_lbls = list(vars_ws.values)[0]
    c_time_col = vars_lbls.index('c_start')+1
    r_time_col = c_time_col + 1
    c_stress_col = c_time_col + 2
    r_stress_col = c_time_col + 3
    marker_vars = [('c_start', c_time_col, c_stress_col),('r_start', r_time_col, r_stress_col)]
    for ser_tit, time_col, stress_col in marker_vars:
        #marker data series
        c_ref1 = opyx.chart.Reference(vars_ws, min_col=time_col, min_row=ws_num+1)
        c_ref2 = opyx.chart.Reference(vars_ws, min_col=stress_col, min_row=ws_num+1)
        ser1 = opyx.chart.Series(c_ref2, c_ref1, title=ser_tit)
        #marker properties
        marker_size = 20.0
        marker_width = 50000 #EMU size
        ser1.marker = opyx.chart.marker.Marker(symbol='x', size=marker_size)
        ser1.graphicalProperties.line.noFill = True
        ser1.marker.graphicalProperties.line.width = marker_width
        ser1.marker.graphicalProperties.line.solidFill = opyx.drawing.colors.ColorChoice(prstClr='black')
        #add to charts
        c1.append(ser1)
    
    
    return

##

def chart_default_settings(c1, axes_set_lb='yes', sec_ax='no', x_minor_ticks='no'):
    '''Set default settings for charts: remove gridlines, add major tickmarks, increase axis line width, add line around plot area, remove line around chart area. '''
    
    #line properties
    w_emu_1p = int(10000/0.79) #emu size to get 1p width
    w_emu_1p5 = int(w_emu_1p * 1.5)
    lprop = opyx.drawing.line.LineProperties(solidFill=opyx.drawing.colors.ColorChoice(prstClr='black'), prstDash='solid', w=w_emu_1p5)
    #shape properties
    no_line = opyx.chart.shapes.GraphicalProperties(ln=opyx.drawing.line.LineProperties(noFill=True))
    lprop_black_1p5 = opyx.chart.shapes.GraphicalProperties(ln=lprop)
    #remove major gridlines
    try:
        c1.x_axis.majorGridlines.spPr = no_line
    except AttributeError: #linechart causes error
        pass
    c1.y_axis.majorGridlines.spPr = no_line
    #modify axis lineproperties
    c1.x_axis.spPr = lprop_black_1p5
    c1.x_axis.majorTickMark = 'in'
    if x_minor_ticks == 'yes':
        c1.x_axis.minorTickMark = 'in'
    c1.y_axis.spPr = lprop_black_1p5
    c1.y_axis.majorTickMark = 'in'
    
    c1.plot_area.spPr = lprop_black_1p5

    c1.graphical_properties = no_line #hox! this required a change in _chart.py (check comment at start of this script)
    
    #set axes into left-bot corner
    if axes_set_lb == 'yes':
        c1.x_axis.crosses = 'min'
        c1.y_axis.crosses = 'min'
    #add secondary axis
    if sec_ax == 'yes':
        c1.y_axis.crosses = 'max'
    
    return

##read compression|relaxation excel files into python

def tab_data_from_excel(file_pth, ang_data_chk):
    '''reads x and y values from a workbook (containing gleeble results by Juha) and creates lists out of the values.
        Return values_d'''
    
    #loads an existing workbook. note! data_only=True allows you to read the result of a formula, not the formula itself
    my_wb = opyx.load_workbook(filename=file_pth, data_only=True)
    ws = my_wb.worksheets[0]
    values_d = orddict()
    #labels
    labels1 = list(ws.values)[0]
    labels2 = list(ws.values)[1]
    labels = []
    for i, lbl1 in enumerate(labels1):
        lbl2 = labels2[i]
        try:
            labels.append(lbl1+lbl2)
        except TypeError:
            labels.append(lbl1)
    #values
    values = []
    all_cols = list(ws.columns)
    get_value = lambda x:x.value
    for col in all_cols:
        col = col[2:]
        vals = list(map(get_value, col))
        values.append(vals)
    #combine
    for i, lbl in enumerate(labels):
        if lbl == None: #skip empty lines in data workbook
            continue
        values_d[lbl] = values[i]
    
    return values_d

def tab_data_from_excel_read_only(file_pth): #seems to be bad, no remarkable speed-up!
    ''' '''
    
    #loads an existing workbook. note! data_only=True allows you to read the result of a formula, not the formula itself
    my_wb = opyx.load_workbook(filename=file_pth, data_only=True, read_only=True)
    ws = my_wb.worksheets[0]
    values_d = orddict()
    #labels
    labels1 = list(ws.values)[0]
    labels2 = list(ws.values)[1]
    labels = []
    for i, lbl1 in enumerate(labels1):
        lbl2 = labels2[i]
        try:
            labels.append(lbl1+lbl2)
        except TypeError:
            labels.append(lbl1)
    #values
    values = []
    for i in range(len(labels)):
        values.append([])
    for row in ws.rows:
        for i, cell in enumerate(row):
            try:
                val = float(cell.value)
            except (ValueError, TypeError):
                continue
            values[i].append(val)
    #combine
    for i, lbl in enumerate(labels):
        if lbl == None:
            continue
        values_d[lbl] = values[i]
    
    my_wb.close() #remember to do this always for read_only workbooks
    
    return values_d

def read_csv(file_pth, headers='', header_type='double', convert_type='float'):
    ''' '''
    
    dat_d = orddict()
    #get file data from csv-type data
    with open(file_pth, mode='r', encoding='utf-8-sig') as csvfile:
        if headers == '':
            heads = []
            heads1_str = csvfile.readline()
            if '\t' in heads1_str:
                splitter = '\t'
            elif ';' in heads1_str:
                splitter = ';'
            else:
                splitter = ','
            heads1 = heads1_str.rstrip().split(splitter)
            if header_type == 'double':
                heads2_str = csvfile.readline()
                heads2 = heads2_str.rstrip().split(splitter)
                for i, head1 in enumerate(heads1):
                    try:
                        head2 = heads2[i]
                    except IndexError:
                        head2 = ''
                    heads.append('%s%s' %(head1, head2))
            elif header_type == 'single':
                heads = heads1
        else:
            heads = headers
            splitter = '\t'
            
        for line in csvfile:
            if line == '':
                continue
            line = line.rstrip().lstrip().split(splitter)
            #conversion options
            if convert_type == 'float':
                line1 = []
                for val in line:
                    try:
                        val = float(val)
                    except ValueError: #prevent error for short columns that end before end-of-file
                        val = ''
                    line1.append(val)
            elif convert_type == 'none':
                line1 = line
            # line1 = list(map(float, line))
            for i, head in enumerate(heads):
                data_p = line1[i]
                try:
                    dat_d[head].append(data_p)
                except KeyError:
                    dat_d[head] = [data_p]
                
    
    return dat_d

if __name__ == '__main__':
    
    
    
    pass